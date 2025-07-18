# django imports
from django.views import View
from django.http import JsonResponse, HttpResponse
from django.utils import timezone

# local imports
from core.mixins import LoginRequiredMixin
from contratos.models import Contrato

# third-party imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO



class GenerarReporteView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        try:
            columnas_seleccionadas = request.POST.getlist('columnas_seleccionadas')
            nombre_reporte = request.POST.get('nombre_reporte', 'Reporte_Contratos')

            if not columnas_seleccionadas:
                return JsonResponse({'error': 'Debe seleccionar al menos una columna'}, status=400)

            columnas_disponibles = {
                'ID': 'id',
                'Cliente': 'cliente',
                'Numero Idetificación': 'cliente__numero_documento',
                'Placa Vehículo': 'carro__placa',
                'Marca': 'carro__marca',
                'Modelo': 'carro__modelo',
                'Fecha Inicio Contrato': 'fecha_inicio',
                'Estado': 'activo'
            }

            columnas_validas = [col for col in columnas_seleccionadas if col in columnas_disponibles]

            if not columnas_validas:
                return JsonResponse({'error': 'Columnas seleccionadas no válidas'}, status=400)

            contratos = Contrato.objects.select_related('cliente', 'carro').all().order_by('-fecha_inicio')

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte de Contratos"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_index, columna in enumerate(columnas_validas, 1):
                cell = ws.cell(row=1, column=col_index, value=columna)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for row_index, contrato in enumerate(contratos, 2):
                for col_index, columna in enumerate(columnas_validas, 1):
                    if columna == 'ID':
                        value = contrato.id
                    elif columna == 'Cliente':
                        value = str(contrato.cliente.get_full_name())
                    elif columna == 'Numero Idetificación':
                        value = contrato.cliente.numero_documento
                    elif columna == 'Placa Vehículo':
                        value = contrato.carro.placa
                    elif columna == 'Marca':
                        value = contrato.carro.marca
                    elif columna == 'Modelo':
                        value = contrato.carro.modelo
                    elif columna == 'Fecha Inicio Contrato':
                        value = contrato.fecha_inicio.strftime('%Y-%m-%d')
                    elif columna == 'Estado':
                        value = 'Activo' if contrato.activo else 'Inactivo'
                    else:
                        value = ''
                    ws.cell(row=row_index, column=col_index, value=value)

            for col_index in range(1, len(columnas_validas) + 1):
                column_letter = get_column_letter(col_index)
                ws.column_dimensions[column_letter].width = 15

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{nombre_reporte}_{timestamp}.xlsx"

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            return response

        except Exception as e:
            return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)


